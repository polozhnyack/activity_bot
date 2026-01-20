from aiogram import Router
from aiogram_dialog import DialogManager
from aiogram.types import Message, CallbackQuery, Chat, FSInputFile
from aiogram.types import ContentType, ReplyKeyboardMarkup, KeyboardButton, KeyboardButtonRequestUser, ReplyKeyboardRemove

from aiogram_dialog.widgets.kbd import Select

from dialogs.states import *
from aiogram_dialog.widgets.input import MessageInput

from models.methods import UserService, ChildService
from models.models import *
from logger import logger

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy import select
from datetime import datetime

from utils import delete_file


async def creator_child(callback: CallbackQuery, button, dialog_manager: DialogManager):
    params = callback.data.split("_")
    param = params[0]
    
    logger.debug(param)

    dialog_manager.dialog_data["param"] = param
    await dialog_manager.switch_to(state=AdminState.child_create_or_delete)


async def child_handler(
        message: Message,
        message_input: MessageInput,
        dialog_manager: DialogManager,
        **kwargs
):
    param = dialog_manager.dialog_data.get("param")
    child_service: ChildService = dialog_manager.middleware_data["ChildService"]

    input_text = message.text.strip()

    if param == "create":
        parts = input_text.split(",")
        if len(parts) < 2:
            await message.answer("❌ Нужно ввести ФИ и дату рождения через запятую (например: Иван Иванов, 11.11.2011)")
            return
        full_name = parts[0].strip()
        from datetime import datetime
        try:
            birth_date = datetime.strptime(parts[1].strip(), "%d.%m.%Y").date()
        except ValueError:
            await message.answer("❌ Неверный формат даты. Используйте ДД.MM.ГГГГ")
            return
        
        dialog_manager.dialog_data.update(
            full_name=full_name,
            birth_date=birth_date
        )


        await dialog_manager.switch_to(AdminState.select_child_level)

    elif param == "delete":
        child_code = input_text
        success = await child_service.delete_by_code(child_code)
        if success:
            await message.answer(f"✅ Ребёнок {child_code} удалён")
        else:
            await message.answer(f"❌ Ребёнок {child_code} не найден")

        await dialog_manager.switch_to(AdminState.admin_menu)



async def on_level_selected(
    callback: CallbackQuery,
    widget,
    dialog_manager: DialogManager,
    item_id: str
):
    level_id = int(item_id)
    dialog_manager.dialog_data["level_id"] = level_id

    full_name = dialog_manager.dialog_data.get("full_name")
    birth_date = dialog_manager.dialog_data.get("birth_date")

    child_service: ChildService = dialog_manager.middleware_data["ChildService"]

    child = await child_service.create(
        full_name=full_name,
        birth_date=birth_date,
        level_id=level_id
    )

    await callback.message.answer(f"✅ Ребёнок {child.full_name} ({child.code}) создан")
    await dialog_manager.start(AdminState.admin_menu)




async def user_contact_handler(
        message: Message,
        message_input: MessageInput,
        dialog_manager: DialogManager,
        **kwargs
):
    
    logger.debug("Запущен user_contact_handler")
    user_service: UserService = dialog_manager.middleware_data["UserService"]

    if message.user_shared:
        user_id = message.user_shared.user_id
        selected_role = dialog_manager.dialog_data.pop("selected_role", None)

        if not selected_role:
            await message.answer("❌ Роль не выбрана.")
            return

        updated = await user_service.update_role(
            user_id=user_id,
            new_role_str=selected_role
        )

        if updated:
            chat: Chat = await message.bot.get_chat(user_id)
            full_name = getattr(chat, "full_name", f"ID {user_id}")

            role_enum = UserRole(selected_role.lower())
            role_title = ROLE_NAMES.get(role_enum, selected_role)
            await message.answer(
                f"✅ Пользователю {full_name} успешно присвоена роль: {role_title}"
            )
        else:
            await message.answer(
                f"❌ Не удалось изменить роль пользователя с ID {user_id}. Возможно, роль некорректна или пользователь не найден."
            )
    else:
        await message.answer("❌ Не удалось получить информацию о пользователе.")

    await dialog_manager.switch_to(AdminState.admin_menu)


keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [
            KeyboardButton(
                text="Отправить пользователя 👀",
                request_user=KeyboardButtonRequestUser(
                    request_id=123,
                    user_is_bot=False
                )
            )
        ]
    ],
    resize_keyboard=True,
    one_time_keyboard=True,
)


async def on_role_selected(
    callback: CallbackQuery,
    widget: Select,
    manager: DialogManager,
    item_id: str,
):
    
    message = await callback.bot.send_message(
        chat_id=callback.from_user.id,
        text="📁",
        reply_markup=keyboard
    )
    
    manager.dialog_data["pupa"] = message.message_id
    manager.dialog_data["selected_role"] = item_id

    await callback.answer(f"✅ Вы выбрали роль: {item_id}")
    await manager.switch_to(AdminState.user_select)


async def go_back_admin_menu(callback: CallbackQuery, button, dialog_manager: DialogManager):
    pupa = dialog_manager.dialog_data.get("pupa")
    if pupa:
        try:
            message = await callback.message.answer(
                text=callback.message.text or "🔄",
                reply_markup=ReplyKeyboardRemove()
            )

            await message.delete()

            await callback.bot.delete_message(
                chat_id=callback.from_user.id,
                message_id=pupa
            )

        except Exception as e:
            message = await callback.message.answer(
                "🔄",
                reply_markup=ReplyKeyboardRemove()
            )
            await message.delete()

    dialog_manager.dialog_data.clear()
    await dialog_manager.switch_to(state=AdminState.admin_menu)





async def export_children_to_excel(callback: CallbackQuery, button, dialog_manager: DialogManager):
    """Экспортирует всех детей в Excel с полями code, full_name, birth_date."""
    session = dialog_manager.middleware_data["session"]
    userService: UserService = dialog_manager.middleware_data["UserService"]

    result = await session.execute(
        select(Child.code, Child.full_name, Child.birth_date)
    )
    children = result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Дети"

    headers = ["Код ребёнка", "ФИО", "Дата рождения"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 25

    for code, full_name, birth_date in children:
        ws.append([
            code,
            full_name or "—",
            birth_date.strftime("%d.%m.%Y") if birth_date else "—"
        ])

    center_alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = center_alignment

    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        adjusted_width = length + 2
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    filename = f"Таблица_кодов_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
    wb.save(filename)

    user: User = await userService.get_by_id(callback.from_user.id)

    if user.role == UserRole.admin:
        await callback.bot.send_document(
            chat_id=callback.from_user.id,
            document=FSInputFile(path=filename),
            caption=f"<b>Таблица кодов всех детей</b>",
            parse_mode="HTML"
        )

        delete_file(filename)
    return 
